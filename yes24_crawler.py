from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import pandas as pd
import time
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Chrome 옵션 설정
chrome_options = Options()
chrome_options.add_argument('--headless')
chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument('--disable-dev-shm-usage')

# WebDriver 설정
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=chrome_options)

# YES24 베스트셀러 기본 URL
base_url = "https://www.yes24.com/product/category/bestseller?categoryNumber=001&pageNumber={}&pageSize=24"

book_list = []

# 페이지 범위 설정
for page in range(1, 21):
    url = base_url.format(page)
    driver.get(url)
    time.sleep(2)

    items = driver.find_elements(By.CSS_SELECTOR, 'div.itemUnit')

    if not items:
        print(f"⚠️ 페이지 {page}에서 책 아이템을 찾을 수 없습니다. 크롤링을 중단합니다.")
        break

    for idx, item in enumerate(items):
        try:
            rank = (page - 1) * 24 + idx + 1
            title = item.find_element(By.CLASS_NAME, 'gd_name').text.replace('\n', ' ').strip()

            author = ""
            publisher = ""

            try:
                author_element = item.find_element(By.CLASS_NAME, 'info_auth')
                author = author_element.text.strip()
            except:
                pass

            try:
                publisher_element = item.find_element(By.CLASS_NAME, 'info_pub')
                publisher = publisher_element.text.strip()
            except:
                pass

            book_list.append({
                '순위': rank,
                '책제목': title,
                '저자': author,
                '출판사': publisher
            })

        except Exception as e:
            print(f"❌ 오류 발생 (페이지 {page}, 책 {idx + 1}): {e}")
            continue

driver.quit()

# 리스트를 DataFrame으로 변환
df = pd.DataFrame(book_list)

# 엑셀 파일로 저장
excel_file_name = "yes24_bestsellers_all.xlsx"
df.to_excel(excel_file_name, index=False, engine='openpyxl')

print(f"✅ 크롤링 완료! {excel_file_name} 저장됨. 이제 엑셀 서식을 적용합니다.")

# openpyxl을 사용하여 엑셀 파일 서식 적용
try:
    wb = load_workbook(excel_file_name)
    ws = wb.active

    # 열 너비 자동 조절
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        if column_letter == 'B': # '책제목' 열 (B열)은 더 넓게 설정
            ws.column_dimensions[column_letter].width = 60
        else:
            ws.column_dimensions[column_letter].width = adjusted_width

    # '책제목' 열에 텍스트 줄 바꿈 적용
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrapText=True)

    wb.save(excel_file_name)
    print(f"✅ {excel_file_name} 파일에 엑셀 서식이 적용되었습니다.")

except Exception as e:
    print(f"❌ 엑셀 서식 적용 중 오류 발생: {e}")
